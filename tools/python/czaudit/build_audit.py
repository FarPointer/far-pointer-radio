"""Compare the OneNote set lists against Spinitron's logged spins and write an audit workbook.

The two sources are not equally authoritative, and the workbook is shaped around that:

  * OneNote notes are show *prep*, written before air. A track listed there may never
    have been played, and the note may stop partway through the episode.
  * Spinitron is the as-aired log, but it is what we are auditing, so it can be wrong
    in both directions.

So a track in the notes with no spin is a *candidate* to add, not a proven omission,
and a spin with no note entry is usually just an unfinished note rather than a bad
log. Each episode therefore gets a coverage verdict that says how far its note can be
trusted, and the spins-without-notes sheet is explicitly not a removal list.
"""
import csv
import datetime as dt
import json
import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from matching import artist_score, norm, ratio

HERE = Path(__file__).parent
SOURCES = (Path(__file__).resolve().parents[3]
           / "shows/convergence-zone/playlists/sources")
SPINS_CSV = SOURCES / "spinitron" / "Spinssearchresults84208326forKSER.csv"
EPISODES_JSON = HERE / "onenote_episodes.json"
SITE_JSON = HERE / "site_episodes.json"
OUT_XLSX = HERE / "cz-playlist-spinitron-audit.xlsx"

def parse_elapsed(v: str):
    """'02:50' or '01:11:03' -> timedelta. Returns None if unparseable."""
    v = (v or "").strip()
    if not re.fullmatch(r"\d{1,2}(:\d{2}){1,2}", v):
        return None
    parts = [int(p) for p in v.split(":")]
    if len(parts) == 2:
        return dt.timedelta(minutes=parts[0], seconds=parts[1])
    return dt.timedelta(hours=parts[0], minutes=parts[1], seconds=parts[2])


def fmt_clock(t) -> str:
    return t.strftime("%-I:%M %p").lower() if t else ""


# ---------------------------------------------------------------- load sources

def load_spins():
    by_date = {}
    with open(SPINS_CSV, encoding="utf-8-sig") as fh:
        for r in csv.DictReader(fh):
            d = dt.datetime.strptime(r["Playlist Date"], "%b %d, %Y").date()
            when = None
            raw = (r.get("Date-time") or "").strip()
            if raw:
                try:
                    when = dt.datetime.fromisoformat(raw)
                except ValueError:
                    when = None
            by_date.setdefault(d, []).append({
                "artist": (r.get("Artist") or "").strip(),
                "song": (r.get("Song") or "").strip(),
                "release": (r.get("Release") or "").strip(),
                "label": (r.get("Label") or "").strip(),
                "when": when,
                "start": (r.get("Playlist Time") or "").strip(),
            })
    for d in by_date:
        by_date[d].sort(key=lambda s: (s["when"] is None, s["when"]))
        for i, s in enumerate(by_date[d], 1):
            s["idx"] = i
    return by_date


# ---------------------------------------------------------------- matching

def match_episode(tracks, spins):
    """Greedy multi-pass match. Returns (matches, note_only, spin_only)."""
    for t in tracks:
        t["_a"], t["_s"] = norm(t["artist"]), norm(t["song"])
        t["_a2"], t["_s2"] = norm(t["artist"], True), norm(t["song"], True)
    for s in spins:
        s["_a"], s["_s"] = norm(s["artist"]), norm(s["song"])
        s["_a2"], s["_s2"] = norm(s["artist"], True), norm(s["song"], True)

    matches, used_t, used_s = [], set(), set()

    def take(ti, si, conf, why):
        used_t.add(ti)
        used_s.add(si)
        matches.append({"track": tracks[ti], "spin": spins[si],
                        "confidence": conf, "basis": why})

    # Pass 1 - artist and song both identical after normalisation.
    for ti, t in enumerate(tracks):
        if ti in used_t:
            continue
        for si, s in enumerate(spins):
            if si in used_s:
                continue
            if t["_a"] == s["_a"] and t["_s"] == s["_s"]:
                take(ti, si, "exact", "artist + song match exactly")
                break

    # Pass 2 - same song, artist spelled differently.
    for ti, t in enumerate(tracks):
        if ti in used_t:
            continue
        best = None
        for si, s in enumerate(spins):
            if si in used_s:
                continue
            if t["_s2"] and t["_s2"] == s["_s2"]:
                r = artist_score(t["_a2"], s["_a2"])
                if r >= 0.60 and (best is None or r > best[1]):
                    best = (si, r)
        if best:
            take(ti, best[0], "strong", f"song matches; artist similarity {best[1]:.2f}")

    # Pass 3 - both fields close but not identical.
    for ti, t in enumerate(tracks):
        if ti in used_t:
            continue
        best = None
        for si, s in enumerate(spins):
            if si in used_s:
                continue
            rs, ra = ratio(t["_s2"], s["_s2"]), artist_score(t["_a2"], s["_a2"])
            if rs >= 0.82 and ra >= 0.60:
                score = rs + ra
                if best is None or score > best[1]:
                    best = (si, score, rs, ra)
        if best:
            take(ti, best[0], "likely",
                 f"song similarity {best[2]:.2f}, artist similarity {best[3]:.2f}")

    # Pass 4 - same artist, and the song is a near miss (often a subtitle differs).
    for ti, t in enumerate(tracks):
        if ti in used_t:
            continue
        best = None
        for si, s in enumerate(spins):
            if si in used_s:
                continue
            if t["_a2"] and artist_score(t["_a2"], s["_a2"]) >= 0.90:
                r = ratio(t["_s2"], s["_s2"])
                if r >= 0.55 and (best is None or r > best[1]):
                    best = (si, r)
        if best:
            take(ti, best[0], "possible",
                 f"same artist; song similarity {best[1]:.2f}")

    note_only = [t for i, t in enumerate(tracks) if i not in used_t]
    spin_only = [s for i, s in enumerate(spins) if i not in used_s]
    return matches, note_only, spin_only


def suggest_time(track, matches, spins, playlist_start):
    """Best guess at where an unlogged track belongs, as (clock string, rationale)."""
    elapsed = parse_elapsed(track.get("elapsed", ""))
    if elapsed is not None and playlist_start is not None:
        return fmt_clock(playlist_start + elapsed), \
            f"note offset {track['elapsed']} from show start"

    # No usable offset: sit it between the matched neighbours on either side.
    before = [m for m in matches if m["track"]["seq"] < track["seq"] and m["spin"]["when"]]
    after = [m for m in matches if m["track"]["seq"] > track["seq"] and m["spin"]["when"]]
    if before:
        prev = max(before, key=lambda m: m["track"]["seq"])
        return "", (f"after {prev['spin']['artist']} - {prev['spin']['song']} "
                    f"({fmt_clock(prev['spin']['when'])})")
    if after:
        nxt = min(after, key=lambda m: m["track"]["seq"])
        return "", (f"before {nxt['spin']['artist']} - {nxt['spin']['song']} "
                    f"({fmt_clock(nxt['spin']['when'])})")
    return "", "position unclear - no matched neighbours"


# ---------------------------------------------------------------- workbook

HEAD_FILL = PatternFill("solid", fgColor="1F3864")
HEAD_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=14, color="1F3864")
NOTE_FONT = Font(italic=True, size=10, color="444444")
AMBER = PatternFill("solid", fgColor="FFF2CC")
GREY = PatternFill("solid", fgColor="EDEDED")
GREEN = PatternFill("solid", fgColor="E2EFDA")
RED = PatternFill("solid", fgColor="FCE4E4")
THIN = Border(*[Side(style="thin", color="BFBFBF")] * 4)


def write_sheet(ws, title, blurb, headers, rows, widths, fills=None, wrap_cols=()):
    ws["A1"] = title
    ws["A1"].font = TITLE_FONT
    ws["A2"] = blurb
    ws["A2"].font = NOTE_FONT
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max(len(headers), 2))
    ws.row_dimensions[2].height = 30
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")

    hr = 4
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=hr, column=c, value=h)
        cell.fill, cell.font, cell.border = HEAD_FILL, HEAD_FONT, THIN
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.row_dimensions[hr].height = 28

    for ri, row in enumerate(rows, hr + 1):
        fill = fills[ri - hr - 1] if fills else None
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.border = THIN
            cell.alignment = Alignment(
                vertical="top", wrap_text=headers[ci - 1] in wrap_cols)
            if fill:
                cell.fill = fill

    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = ws.cell(row=hr + 1, column=1)
    if rows:
        ws.auto_filter.ref = (f"A{hr}:{get_column_letter(len(headers))}"
                              f"{hr + len(rows)}")


def load_reference():
    """Per broadcast, the best available account of what was meant to air.

    Keyed by broadcast date, not by episode, because the show is replayed from
    automation during vacations and hiatus. One published playlist can therefore be the
    reference for several broadcasts -- its original airing and every rerun -- and each
    of those airings is logged separately in Spinitron, so each needs auditing
    separately. scrape_site.py decides which playlist a broadcast is, by song overlap.

    Where convergencezone.fm published a playlist, that is preferred: it is the same
    table as the OneNote note (verified - identical track sets on every episode where
    both exist) but with cleaner text, real accents, and the note's scratch artifacts
    removed. It is therefore a better thing to copy into Spinitron, but it is NOT
    independent evidence that a track aired, since it descends from the same prep
    document.
    """
    if not EPISODES_JSON.exists():
        raise SystemExit(f"missing {EPISODES_JSON.name} - run extract.py first")
    notes = json.loads(EPISODES_JSON.read_text(encoding="utf-8"))
    # Step 2 needs downloaded HTML that is not kept in the repo, so treat the site
    # scrape as optional: without it every episode falls back to its OneNote note.
    if SITE_JSON.exists():
        site = json.loads(SITE_JSON.read_text(encoding="utf-8"))
    else:
        print(f"note: {SITE_JSON.name} absent - using OneNote notes for every episode")
        site = {"playlists": {}, "broadcasts": {}}

    ref = {}
    for date, tracks in notes.items():
        ref[date] = {
            "tracks": tracks, "source": "OneNote prep note", "url": "",
            "file": tracks[0]["source_file"] if tracks else "",
            "replay_of": "", "bind_score": "",
        }
    for date, b in site.get("broadcasts", {}).items():
        p = site["playlists"][b["slug"]]
        if not p["tracks"]:
            continue
        first = p.get("first_aired", "")
        ref[date] = {
            "tracks": [{
                "seq": t["position"], "artist": t["artist"], "song": t["song"],
                "album": t["album"], "label": t.get("label", ""),
                "elapsed": t.get("time", ""), "notes": t.get("notes", "")[:400],
                "source_file": b["slug"],
            } for t in p["tracks"]],
            "source": "convergencezone.fm published playlist",
            "url": p.get("url", ""), "file": b["slug"],
            # A rerun is audited against the playlist of the episode it repeats.
            "replay_of": first if first and first != date else "",
            "bind_score": b.get("score", ""),
        }
    return ref


def main():
    reference = load_reference()
    episodes = {d: r["tracks"] for d, r in reference.items()}
    spins_by_date = load_spins()

    summary, add_rows, add_fills = [], [], []
    meta_rows, meta_fills, extra_rows, extra_fills = [], [], [], []

    for date_str in sorted(episodes):
        d = dt.date.fromisoformat(date_str)
        tracks = episodes[date_str]
        spins = spins_by_date.get(d, [])

        playlist_start = None
        if spins and spins[0]["start"]:
            try:
                tod = dt.datetime.strptime(spins[0]["start"], "%I:%M:%S %p").time()
                playlist_start = dt.datetime.combine(d, tod)
            except ValueError:
                playlist_start = None

        matches, note_only, spin_only = match_episode(tracks, spins)
        cov = len(matches) / len(spins) if spins else 0.0

        # Two ratios, and they mean different things. Coverage says how much of the
        # Spinitron log the note accounts for; logged says how much of the note made
        # it into Spinitron. High coverage with low logged is the actionable case:
        # the note tracks the show well and Spinitron is the one missing entries.
        logged = len(matches) / len(tracks) if tracks else 0.0
        if not spins:
            verdict = "no Spinitron playlist"
        elif cov < 0.55:
            verdict = "note is PARTIAL - unmatched spins are expected, not errors"
        elif logged >= 0.85:
            verdict = "both lists agree closely - little to do"
        elif logged >= 0.60:
            verdict = "Spinitron looks short a few spins"
        else:
            verdict = "Spinitron appears to be missing much of this show"

        ref = reference[date_str]
        summary.append([
            date_str, "replay" if ref["replay_of"] else "original",
            ref["replay_of"], ref["source"], ref["file"],
            len(tracks), len(spins), len(matches),
            len(note_only), len(spin_only),
            f"{cov:.0%}", f"{logged:.0%}", verdict,
        ])

        for t in note_only:
            when, why = suggest_time(t, matches, spins, playlist_start)
            confident = bool(when)
            # If the same artist has an unmatched spin in this episode, the planned
            # track was most likely swapped for another rather than dropped from the
            # log. Editing that spin beats adding a second one.
            sub = ""
            for s in spin_only:
                if artist_score(norm(t["artist"], True), norm(s["artist"], True)) >= 0.90:
                    sub = (f"Spinitron logged '{s['song']}' by this artist at "
                           f"{fmt_clock(s['when'])} - may be a substitution; "
                           f"edit that spin instead of adding one")
                    break
            if sub:
                assess = "possible substitution - see next column"
            elif confident:
                assess = "likely omission"
            else:
                assess = "check - position uncertain"
            add_rows.append([
                date_str, when, t["artist"], t["song"], t["album"], t["label"],
                assess, sub, why, t["elapsed"], reference[date_str]["source"],
                t["notes"][:200], t["source_file"],
            ])
            add_fills.append(AMBER if sub else (GREY if not confident else None))

        for m in matches:
            t, s = m["track"], m["spin"]
            diffs = []
            if norm(t["artist"]) != norm(s["artist"]):
                diffs.append(f"artist: note '{t['artist']}' vs Spinitron '{s['artist']}'")
            if norm(t["song"]) != norm(s["song"]):
                diffs.append(f"song: note '{t['song']}' vs Spinitron '{s['song']}'")
            if t["album"] and s["release"] and norm(t["album"]) != norm(s["release"]):
                diffs.append(f"album: note '{t['album']}' vs Spinitron '{s['release']}'")
            if not diffs:
                continue
            meta_rows.append([
                date_str, fmt_clock(s["when"]), s["artist"], s["song"], s["release"],
                t["artist"], t["song"], t["album"], m["confidence"],
                m["basis"], " | ".join(diffs),
            ])
            meta_fills.append(AMBER if m["confidence"] in ("likely", "possible") else None)

        partial = cov < 0.55
        for s in spin_only:
            extra_rows.append([
                date_str, fmt_clock(s["when"]), s["artist"], s["song"], s["release"],
                s["label"],
                "note incomplete - almost certainly fine"
                if partial else "not in the note - worth a look",
                verdict,
            ])
            extra_fills.append(GREY if partial else AMBER)

    wb = Workbook()

    ws = wb.active
    ws.title = "Summary"
    write_sheet(
        ws, "Convergence Zone - published playlists vs Spinitron audit",
        "One row per BROADCAST that has both a set list and a Spinitron playlist. "
        "Replays count separately: an episode re-aired from automation is logged again "
        "in Spinitron, so it is audited again here against the same playlist - see the "
        "'Airing' and 'Repeat of' columns. 'Coverage' = share of Spinitron spins the "
        "reference accounts for; low coverage means the reference is partial, not that "
        "Spinitron is wrong. 'Reference logged' = share of reference tracks that reached "
        "Spinitron; low values there are the ones worth acting on. Read Method first.",
        ["Broadcast date", "Airing", "Repeat of", "Reference used", "Source",
         "Tracks in reference", "Spins logged", "Matched", "Missing from Spinitron",
         "In Spinitron only", "Coverage", "Reference logged", "Verdict"],
        summary,
        [15, 10, 12, 34, 40, 15, 12, 10, 16, 16, 10, 15, 50],
        fills=[GREEN if r[12].startswith("both lists agree")
               else (GREY if r[12].startswith("note is PARTIAL")
                     else (RED if r[12].startswith("Spinitron appears") else AMBER))
               for r in summary],
        wrap_cols={"Verdict", "Reference used", "Source"},
    )

    write_sheet(
        wb.create_sheet("Add to Spinitron"),
        "Tracks in the notes with no matching spin",
        "These are the strongest candidates for spins missing from Spinitron. They are "
        "candidates, not proof: the notes were written before air, so a track here may "
        "simply never have been played. Grey rows are ones whose position could not be "
        "pinned down; amber rows are ones where the same artist has an unlogged spin, so "
        "the track was probably swapped rather than dropped - edit that spin instead "
        "of adding a new one. Suggested time = show start plus the note's offset.",
        ["Air date", "Add at", "Artist", "Song", "Album", "Label", "Assessment",
         "Possible substitution", "Where it fits", "Offset", "Reference used",
         "Context", "Source"],
        add_rows, [11, 10, 28, 36, 28, 16, 26, 50, 36, 9, 34, 40, 24],
        fills=add_fills,
        wrap_cols={"Possible substitution", "Where it fits", "Context",
                   "Song", "Album", "Assessment", "Reference used"},
    )

    write_sheet(
        wb.create_sheet("Check metadata"),
        "Matched spins whose details disagree with the note",
        "The spin exists in Spinitron, but the artist, song, or album text differs from "
        "the note. Most are harmless spelling or punctuation variants; amber rows were "
        "matched on similarity rather than an exact hit, so confirm the pairing before "
        "editing anything.",
        ["Air date", "Spin time", "Spinitron artist", "Spinitron song",
         "Spinitron release", "Note artist", "Note song", "Note album",
         "Match confidence", "Matched on", "What differs"],
        meta_rows, [11, 10, 24, 30, 26, 24, 30, 24, 15, 30, 54],
        fills=meta_fills,
        wrap_cols={"What differs", "Matched on", "Spinitron song", "Note song"},
    )

    write_sheet(
        wb.create_sheet("Spins not in notes"),
        "Spins with no counterpart in the note - NOT a removal list",
        "Every one of these is logged in Spinitron but absent from the note. For "
        "episodes whose note is partial (grey rows) that is expected and needs no "
        "action. Amber rows come from episodes whose note otherwise tracks the show "
        "closely, so they are the only ones worth a second look - and even then the "
        "likeliest explanation is a last-minute substitution.",
        ["Air date", "Spin time", "Artist", "Song", "Release", "Label",
         "Assessment", "Episode verdict"],
        extra_rows, [11, 10, 26, 34, 28, 22, 34, 46],
        fills=extra_fills, wrap_cols={"Assessment", "Episode verdict", "Song"},
    )

    method_rows = [
        ["What this compares", "Spinitron's as-aired log against the best available "
                               "account of what each show was meant to contain."],
        ["Reference used", "Where a convergencezone.fm playlist matches the broadcast "
                           "that is the reference; otherwise the OneNote prep note. The "
                           "Summary sheet names the one used per row."],
        ["Replays are audited too", "The show is replayed from automation during "
                                    "vacations and hiatus. Each replay is logged in "
                                    "Spinitron as its own playlist, so each is audited "
                                    "as its own row against the playlist of the episode "
                                    "it repeats. See the 'Airing' and 'Repeat of' "
                                    "columns on Summary."],
        ["How a broadcast is identified", "By songs, not dates. A page states the date "
                                          "it was written for, which for a replay is "
                                          "months before the airing being audited. Each "
                                          "Spinitron playlist is matched to the "
                                          "published playlist it overlaps most. "
                                          "Distinct consecutive weeks overlap about "
                                          "0.00-0.08 and genuine replays 0.6-0.96, so "
                                          "the two are easy to tell apart; a replay "
                                          "claim below 0.60 is rejected."],
        ["IMPORTANT - not independent", "The published playlist and the OneNote table "
                                        "are the SAME document: track sets are "
                                        "identical on every episode where both exist. "
                                        "The site therefore does NOT independently "
                                        "confirm a track aired. It is preferred only "
                                        "because its text is cleaner."],
        ["What a replay row means", "A replay plays the recording of the original show, "
                                    "so its true content is what that episode ACTUALLY "
                                    "aired - which may differ from the published plan. "
                                    "The plan is still the best reference available, "
                                    "but a missing track on a replay row can mean the "
                                    "original never played it either."],
        ["Why prefer the site text", "It carries real accents (Sinead -> Sinead with "
                                     "acute, Sigur Ros, Eydis), fixes note typos "
                                     "('MIDI Jani tor', 'C ity of Mirrors', 'Da ncing', "
                                     "'Remis'), and drops scratch artifacts like "
                                     "'MIC BREAK brass clouds'. It is the better thing "
                                     "to copy into Spinitron - but not uniformly: the "
                                     "site spells 'Erwillian' where the note has the "
                                     "correct 'Erwilian'."],
        ["Spinitron export", f"{SPINS_CSV.name} - 3,282 spins across 164 playlists, "
                             f"2023-05-30 to 2026-07-28."],
        ["Broadcasts audited", f"{len(summary)} of 164 Spinitron playlists - the ones "
                               f"with a matching set list. The rest have neither a "
                               f"published playlist nor a usable prep note."],
        ["Not auditable", "67 of 83 OneNote files are promo blurbs with no track "
                          "listing. 39 of 66 site posts are promo prose with no "
                          "playlist table. Four published playlists never matched any "
                          "Spinitron broadcast (episodes 003, 009, 010, 011): they "
                          "aired before Spinitron logging began on 2023-05-30 and were "
                          "never replayed."],
        ["Direction of trust", "Spinitron is the as-aired log; the reference is a plan "
                               "written beforehand. A reference entry with no spin is a "
                               "candidate omission. A spin with no reference entry is "
                               "usually just an unfinished reference."],
        ["Matching", "Four passes, most to least strict: exact artist+song; same song "
                     "with a differently spelled artist; both fields similar; same "
                     "artist with a near-miss song. Comparison ignores case, accents, "
                     "punctuation, leading articles, and 'feat.' credits. Artists match "
                     "when one name's words are a subset of the other's, so "
                     "'Sin Fang' pairs with 'Sin Fang, Kjartan Holm, Fischersund'."],
        ["Known limits", "Times in the reference are offsets from show start, not clock "
                         "times, so suggested times assume the show began when "
                         "Spinitron says it did. Some episodes carry no times at all. "
                         "OneNote column layouts differ per file and were mapped by "
                         "hand."],
        ["Not covered", "MichaelG's .xlsx workbooks - see cz-missing-spins.xlsx for "
                        "those."],
    ]
    write_sheet(
        wb.create_sheet("Method"),
        "How this workbook was built, and what it cannot tell you",
        "Read this before making edits in Spinitron.",
        ["Topic", "Detail"], method_rows, [26, 110], wrap_cols={"Detail"},
    )

    wb.save(OUT_XLSX)

    print(f"{'BROADCAST':<11} {'AIRING':<8} {'REF':>4}{'SPIN':>5}{'MATCH':>6}"
          f"{'ADD':>5}{'EXTRA':>6} {'COV':>5}{'LOGGED':>7}  reference")
    for r in summary:
        print(f"{r[0]:<11} {r[1]:<8} {r[5]:>4}{r[6]:>5}{r[7]:>6}{r[8]:>5}{r[9]:>6} "
              f"{r[10]:>5}{r[11]:>7}  {r[4]}"
              f"{' (repeat of ' + r[2] + ')' if r[2] else ''}")
    print(f"\nAdd candidates : {len(add_rows)}")
    print(f"Metadata diffs : {len(meta_rows)}")
    print(f"Spins w/o notes: {len(extra_rows)}")
    print(f"\nwritten: {OUT_XLSX}")


if __name__ == "__main__":
    main()
