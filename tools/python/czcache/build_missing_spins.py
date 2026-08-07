"""Reconcile MichaelG's workbooks against Spinitron and write the missing-spins workbook.

MichaelG writes his workbook after the show, so unlike a set list it is a record rather
than a plan. That inverts the direction of trust the rest of the audit uses: a workbook
track with no spin is a gap in *Spinitron*, and the workbook is what gets copied across.

Two sheets come out of that:

  * `Add missing` -- workbook tracks Spinitron never logged, each with a suggested clock
    time so it can be typed straight into the log.
  * `Remove or replace` -- spins with no workbook counterpart, which are candidates for
    deletion unless they are really one of the missing tracks mistyped.

Every date with a workbook is audited on every run, so a re-export of the Spinitron
search or a newly added workbook is picked up simply by running this again. Nothing
about which episodes exist is written down here.

Run this, then `enrich_missing_spins.py`, which fills the import-ready metadata columns
and refreshes the CSV export:

    uv run --with openpyxl --with beautifulsoup4 --with lxml --with pyyaml \\
        python build_missing_spins.py
    uv run --with openpyxl python enrich_missing_spins.py
"""
import csv
import datetime as dt
import sys

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

import load_workbooks
from paths import CZAUDIT, SHOW, SPINS_CSV

sys.path.insert(0, str(CZAUDIT))
from build_audit import match_episode  # noqa: E402

ANALYSIS = SHOW / "playlists/analysis"
OUT_XLSX = ANALYSIS / "cz-missing-spins.xlsx"

# A suggested slot narrower than this is not really evidence that the track aired -- the
# log is dense enough there that the show had no room for it. Those rows are greyed so
# MichaelG confirms them rather than importing them blind.
TIGHT_FIT_SECONDS = 4 * 60

TITLE_FONT = Font(bold=True, size=12)
BLURB_FONT = Font(size=10)
BODY_FONT = Font(size=10)
DATE_FONT = Font(bold=True, size=10)
ADD_HEAD_FILL = PatternFill("solid", fgColor="D9EAD3")
REMOVE_HEAD_FILL = PatternFill("solid", fgColor="F4CCCC")
GREY = PatternFill("solid", fgColor="EFEFEF")
AMBER = PatternFill("solid", fgColor="FFF2CC")

ADD_HEADERS = ["Air date", "Add at", "Artist", "Song", "Album", "Release", "Local",
               "Duration", "Label", "Confidence", "Where it fits", "Sheet row"]
ADD_WIDTHS = [11, 8, 28, 44, 36, 10, 7, 10, 26, 26, 30, 9]
REMOVE_HEADERS = ["Air date", "Spin time", "Artist", "Song", "Album", "Release", "Local",
                  "Duration", "Label", "Possible replacement"]
REMOVE_WIDTHS = [11, 10, 30, 42, 36, 10, 7, 10, 26, 60]


def load_spins():
    """The Spinitron export grouped by air date, oldest spin first.

    Each date also carries the playlist's scheduled window, which is what bounds a track
    that belongs before the first logged spin or after the last one.
    """
    by_date = {}
    with open(SPINS_CSV, encoding="utf-8-sig") as fh:
        for r in csv.DictReader(fh):
            raw = (r.get("Date-time") or "").strip()
            if not raw:
                continue
            try:
                when = dt.datetime.fromisoformat(raw)
                start = dt.datetime.fromisoformat(r["Playlist Date-time"])
            except ValueError:
                continue
            try:
                minutes = int(float(r.get("Playlist Duration") or 0))
            except ValueError:
                minutes = 0
            day = by_date.setdefault(start.date(), {"start": start, "end": None,
                                                    "spins": []})
            end = start + dt.timedelta(minutes=minutes)
            # A persona switch mid-show splits one broadcast across two playlists, so the
            # window is the union of them rather than whichever row was read first.
            day["start"] = min(day["start"], start)
            day["end"] = end if day["end"] is None else max(day["end"], end)
            day["spins"].append({
                "artist": (r.get("Artist") or "").strip(),
                "song": (r.get("Song") or "").strip(),
                "album": (r.get("Release") or "").strip(),
                "when": when,
            })
    for day in by_date.values():
        day["spins"].sort(key=lambda s: s["when"])
        # A stable index, because two spins can share a timestamp to the second -- the
        # same artist logged twice at 21:36:17 -- and identifying a spin by its clock
        # time alone would quietly conflate them.
        for i, spin in enumerate(day["spins"]):
            spin["idx"] = i
    return by_date


def windows(tracks, matches, day):
    """Group unmatched tracks into the log gaps they fall in.

    Anchors are the *matched* spins, taken in workbook order: an unmatched spin is a
    removal candidate, so letting it bound a gap would shrink the very window the
    missing track has to fit into. Yields (start, end, kind, tracks) with `start` and
    `end` as datetimes and `kind` naming which end of the show the run sits against.
    """
    anchors = sorted(((m["track"]["seq"], m["spin"]["when"]) for m in matches
                      if m["spin"]["when"]), key=lambda a: a[0])
    matched_seqs = {seq for seq, _ in anchors}
    runs, current = [], []
    for t in sorted(tracks, key=lambda t: t["seq"]):
        if t["seq"] in matched_seqs:
            continue
        if current and t["seq"] != current[-1]["seq"] + 1:
            runs.append(current)
            current = []
        current.append(t)
    if current:
        runs.append(current)

    for run in runs:
        before = [when for seq, when in anchors if seq < run[0]["seq"]]
        after = [when for seq, when in anchors if seq > run[-1]["seq"]]
        if before and after:
            yield before[-1], after[0], "gap", run
        elif after:
            yield min(day["start"], after[0]), after[0], "before", run
        elif before:
            yield before[-1], max(day["end"], before[-1]), "after", run
        else:
            yield day["start"], day["end"], "whole", run


def describe(start, end, kind):
    if kind == "before":
        return f"before first logged spin ({end:%H:%M})"
    if kind == "after":
        return f"after last logged spin ({start:%H:%M})"
    if kind == "whole":
        return "no matched spins - position assumed even across the show"
    minutes = round((end - start).total_seconds() / 60)
    return f"log gap {start:%H:%M}–{end:%H:%M} ({minutes} min)"


def audit_date(date, workbook, day):
    """One broadcast -> (add rows, remove rows). Rows are lists in sheet column order."""
    tracks = [dict(t) for t in workbook["tracks"]]
    matches, track_only, spin_only = match_episode(tracks, [dict(s) for s in day["spins"]])
    matched_idx = {m["spin"]["idx"] for m in matches}

    add_rows, hints = [], {}
    for start, end, kind, run in windows(tracks, matches, day):
        step = (end - start).total_seconds() / (len(run) + 1)
        where = describe(start, end, kind)
        confidence = ("likely aired" if step >= TIGHT_FIT_SECONDS
                      else "tight fit — may not have aired")
        placed = []
        for i, t in enumerate(run, 1):
            at = start + dt.timedelta(seconds=step * i)
            placed.append(t)
            add_rows.append({
                "row": [date, f"{at:%H:%M}", t["artist"], t["song"], t["release"] or None,
                        t["released_raw"] or None, None, None, None,
                        confidence, where, t["sheet_row"]],
                "tight": step < TIGHT_FIT_SECONDS,
            })
        # An unmatched spin inside this window is more likely a mistyped version of one
        # of the tracks the window is missing than a phantom, so it is paired off with
        # them in order and flagged for editing instead of deletion. The interval is
        # half-open at the start so a spin sharing a timestamp with an anchor lands in
        # exactly one window and cannot collect two contradictory hints.
        inside = [s for s in day["spins"]
                  if s["idx"] not in matched_idx and start < s["when"] <= end]
        for spin, track in zip(inside, placed):
            hints[spin["idx"]] = f"may actually be: {track['artist']} — {track['song']}"

    remove_rows = []
    for spin in sorted(spin_only, key=lambda s: (s["when"], s["idx"])):
        remove_rows.append({
            "row": [date, f"{spin['when']:%H:%M}", spin["artist"], spin["song"],
                    spin["album"] or None, None, None, None, None,
                    hints.get(spin["idx"])],
            "hint": spin["idx"] in hints,
        })
    return add_rows, remove_rows


def write_sheet(ws, title, blurb, headers, widths, head_fill, rows, fills):
    ws["A1"] = title
    ws["A1"].font = TITLE_FONT
    ws["A2"] = blurb
    ws["A2"].font = BLURB_FONT
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=c, value=h)
        cell.font = DATE_FONT
        cell.fill = head_fill
    for ri, (row, fill) in enumerate(zip(rows, fills), 5):
        for ci, value in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=value)
            cell.font = DATE_FONT if ci == 1 and value else BODY_FONT
            if fill:
                cell.fill = fill
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=4, column=ci).column_letter].width = w
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:{ws.cell(row=4, column=len(headers)).column_letter}" \
                         f"{max(4, 4 + len(rows))}"


def collapse_dates(rows):
    """Blank a repeated air date so each broadcast reads as one block."""
    seen = None
    for row in rows:
        if row[0] == seen:
            row[0] = None
        else:
            seen = row[0]
    return rows


def main():
    workbooks = load_workbooks.load()
    spins_by_date = load_spins()

    add, add_fills, remove, remove_fills = [], [], [], []
    audited, skipped = [], []
    for date in sorted(workbooks):
        day = spins_by_date.get(dt.date.fromisoformat(date))
        if not day:
            skipped.append(date)
            continue
        a, r = audit_date(date, workbooks[date], day)
        audited.append((date, len(a), len(r)))
        add += [x["row"] for x in a]
        add_fills += [GREY if x["tight"] else None for x in a]
        remove += [x["row"] for x in r]
        remove_fills += [AMBER if x["hint"] else None for x in r]

    wb = Workbook()
    write_sheet(
        wb.active, "Spinitron — missing spins to add (Michael G episodes)",
        "All spreadsheet tracks absent from Spinitron, with a suggested insertion "
        "timestamp. Gray rows sit in tight log windows and may not have actually "
        "aired — Michael can confirm.",
        ADD_HEADERS, ADD_WIDTHS, ADD_HEAD_FILL, collapse_dates(add), add_fills)
    wb.active.title = "Add missing"
    write_sheet(
        wb.create_sheet("Remove or replace"),
        "Spinitron — spins with no spreadsheet counterpart",
        "Candidates for removal. Amber rows include a hint: a nearby missing "
        "spreadsheet track that this spin may actually be — EDIT those in place instead "
        "of removing, and skip the matching row on the Add missing sheet.",
        REMOVE_HEADERS, REMOVE_WIDTHS, REMOVE_HEAD_FILL,
        collapse_dates(remove), remove_fills)

    ANALYSIS.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_XLSX)

    for date, n_add, n_remove in audited:
        print(f"{date}  add {n_add:3d}   remove {n_remove:3d}")
    if skipped:
        print(f"no Spinitron playlist for: {', '.join(skipped)}")
    print(f"\n{len(audited)} broadcasts, {len(add)} to add, {len(remove)} to review")
    print(f"written: {OUT_XLSX}")


if __name__ == "__main__":
    main()
