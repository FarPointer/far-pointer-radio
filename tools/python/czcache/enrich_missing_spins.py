"""Add import-ready metadata columns to the missing-spins workbook and CSV.

Reads `shows/convergence-zone/playlists/analysis/cz-missing-spins.xlsx`, fills in
Release (year), Local (Y), Duration (mm:ss), and Label from the cache where known,
then rewrites the `Remove or replace` CSV export.
"""
import csv
import datetime as dt
import json
from collections import defaultdict

from openpyxl import load_workbook

from paths import BROADCASTS, SHOW, SPINS_CSV

ANALYSIS = SHOW / "playlists/analysis"
MISSING_SPINS_XLSX = ANALYSIS / "cz-missing-spins.xlsx"
REMOVAL_CSV = ANALYSIS / "cz-removal-candidates.csv"


def norm(text):
    return " ".join((text or "").strip().casefold().split())


def fmt_date(value):
    if isinstance(value, dt.datetime):
        return value.date().isoformat()
    if isinstance(value, dt.date):
        return value.isoformat()
    return str(value or "").strip()


def fmt_time(value):
    if isinstance(value, dt.datetime):
        return value.strftime("%H:%M")
    if isinstance(value, dt.time):
        return value.strftime("%H:%M")
    return str(value or "").strip()


def fmt_release_year(released_date):
    if not released_date:
        return ""
    return str(released_date).split("-")[0]


def fmt_duration(seconds):
    if not isinstance(seconds, int):
        return ""
    minutes, secs = divmod(seconds, 60)
    return f"{minutes}:{secs:02d}"


def fmt_duration_text(raw):
    v = str(raw or "").strip()
    if not v:
        return ""
    parts = v.split(":")
    try:
        nums = [int(p) for p in parts]
    except ValueError:
        return ""
    if len(nums) == 2:
        return f"{nums[0]}:{nums[1]:02d}"
    if len(nums) == 3:
        return f"{(nums[0] * 60) + nums[1]}:{nums[2]:02d}"
    return ""


def fmt_logged_time(iso_dt):
    if not iso_dt:
        return ""
    try:
        when = dt.datetime.fromisoformat(iso_dt)
    except ValueError:
        return ""
    return when.strftime("%H:%M")


def local_flag(local):
    return "Y" if local is True else ""


def parse_spinitron_date(text):
    v = str(text or "").strip()
    if not v:
        return ""
    try:
        return dt.datetime.strptime(v, "%b %d, %Y").date().isoformat()
    except ValueError:
        return ""


def parse_spinitron_time(row):
    raw_iso = str(row.get("Date-time") or "").strip()
    if raw_iso:
        try:
            return dt.datetime.fromisoformat(raw_iso).strftime("%H:%M")
        except ValueError:
            pass
    raw_time = str(row.get("Time") or "").strip()
    for pattern in ("%I:%M:%S %p", "%I:%M %p"):
        try:
            return dt.datetime.strptime(raw_time, pattern).strftime("%H:%M")
        except ValueError:
            continue
    return ""


def load_spinitron_lookup():
    lookup = defaultdict(list)
    with open(SPINS_CSV, encoding="utf-8-sig") as fh:
        for row in csv.DictReader(fh):
            date = parse_spinitron_date(row.get("Playlist Date"))
            time = parse_spinitron_time(row)
            artist = row.get("Artist") or ""
            song = row.get("Song") or ""
            key = (date, time, norm(artist), norm(song))
            lookup[key].append({
                "release_year": fmt_release_year((row.get("Released") or "").strip()),
                "local": "Y" if (row.get("Local") or "").strip().upper() == "L" else "",
                "duration": fmt_duration_text(row.get("Duration")),
                "label": (row.get("Label") or "").strip(),
            })
    return lookup


def load_cache_lookups():
    add_lookup = defaultdict(list)
    remove_lookup = defaultdict(list)
    for path in sorted(BROADCASTS.glob("*.json")):
        date = path.stem
        data = json.loads(path.read_text(encoding="utf-8"))
        for spin in data.get("spins", []):
            meta = {
                "release_year": fmt_release_year(spin.get("released_date")),
                "local": local_flag(spin.get("local")),
                "duration": fmt_duration(spin.get("duration_seconds")),
                "label": spin.get("label") or "",
            }
            key = (date, norm(spin.get("artist")), norm(spin.get("song")))
            add_lookup[key].append(meta)
            remove_key = (
                date,
                fmt_logged_time(spin.get("logged_at")),
                norm(spin.get("artist")),
                norm(spin.get("song")),
            )
            remove_lookup[remove_key].append(meta)
    return add_lookup, remove_lookup


def take_meta(lookup, key, counters, scope):
    rows = lookup.get(key) or []
    slot = (scope, key)
    idx = counters[slot]
    counters[slot] += 1
    if idx < len(rows):
        return rows[idx]
    return {"release_year": "", "local": "", "duration": "", "label": ""}


def enrich_add_missing(ws, add_lookup):
    if ws["G4"].value != "Local" or ws["I4"].value != "Label":
        ws.insert_cols(7, amount=3)
    ws["F4"] = "Release"
    ws["G4"] = "Local"
    ws["H4"] = "Duration"
    ws["I4"] = "Label"
    ws["J4"] = "Confidence"
    ws["K4"] = "Where it fits"
    ws["L4"] = "Sheet row"

    counters = defaultdict(int)
    current_date = ""
    for row in range(5, ws.max_row + 1):
        artist = ws[f"C{row}"].value
        song = ws[f"D{row}"].value
        if ws[f"A{row}"].value:
            current_date = fmt_date(ws[f"A{row}"].value)
        if not artist or not song:
            continue
        key = (current_date, norm(artist), norm(song))
        meta = take_meta(add_lookup, key, counters, "add")
        if not ws[f"F{row}"].value:
            ws[f"F{row}"] = meta["release_year"]
        ws[f"G{row}"] = meta["local"]
        ws[f"H{row}"] = meta["duration"]
        ws[f"I{row}"] = meta["label"]


def enrich_remove_or_replace(ws, remove_lookup, remove_spinitron_lookup):
    if ws["G4"].value != "Local" or ws["I4"].value != "Label":
        ws.insert_cols(6, amount=4)
    ws["E4"] = "Album"
    ws["F4"] = "Release"
    ws["G4"] = "Local"
    ws["H4"] = "Duration"
    ws["I4"] = "Label"
    ws["J4"] = "Possible replacement"

    counters = defaultdict(int)
    current_date = ""
    for row in range(5, ws.max_row + 1):
        time = ws[f"B{row}"].value
        artist = ws[f"C{row}"].value
        song = ws[f"D{row}"].value
        if ws[f"A{row}"].value:
            current_date = fmt_date(ws[f"A{row}"].value)
        if not time or not artist or not song:
            continue
        key = (current_date, fmt_time(time), norm(artist), norm(song))
        meta = take_meta(remove_lookup, key, counters, "cache")
        if not any(meta.values()):
            meta = take_meta(remove_spinitron_lookup, key, counters, "spinitron")
        ws[f"F{row}"] = meta["release_year"]
        ws[f"G{row}"] = meta["local"]
        ws[f"H{row}"] = meta["duration"]
        ws[f"I{row}"] = meta["label"]


def write_removal_csv(ws):
    rows = []
    current_date = ""
    for row in range(5, ws.max_row + 1):
        if ws[f"A{row}"].value:
            current_date = fmt_date(ws[f"A{row}"].value)
        if not ws[f"B{row}"].value and not ws[f"C{row}"].value:
            continue
        rows.append({
            "air_date": current_date,
            "time": fmt_time(ws[f"B{row}"].value),
            "artist": ws[f"C{row}"].value or "",
            "song": ws[f"D{row}"].value or "",
            "album": ws[f"E{row}"].value or "",
            "release": ws[f"F{row}"].value or "",
            "local": ws[f"G{row}"].value or "",
            "duration": ws[f"H{row}"].value or "",
            "label": ws[f"I{row}"].value or "",
            "hint": ws[f"J{row}"].value or "",
        })
    with open(REMOVAL_CSV, "w", newline="", encoding="utf-8") as fh:
        w = csv.DictWriter(
            fh,
            fieldnames=[
                "air_date", "time", "artist", "song", "album",
                "release", "local", "duration", "label", "hint",
            ],
            lineterminator="\n",
        )
        w.writeheader()
        w.writerows(rows)


def main():
    add_lookup, remove_lookup = load_cache_lookups()
    remove_spinitron_lookup = load_spinitron_lookup()
    wb = load_workbook(MISSING_SPINS_XLSX)
    enrich_add_missing(wb["Add missing"], add_lookup)
    enrich_remove_or_replace(wb["Remove or replace"], remove_lookup, remove_spinitron_lookup)
    wb.save(MISSING_SPINS_XLSX)
    write_removal_csv(wb["Remove or replace"])
    wb.close()
    print(f"updated {MISSING_SPINS_XLSX}")
    print(f"updated {REMOVAL_CSV}")


if __name__ == "__main__":
    main()
